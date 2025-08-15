"""
Enhanced Report generation utilities with professional styling
"""
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak, KeepTogether
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.graphics.shapes import Drawing, Rect, Line
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics import renderPDF
from reportlab.lib.colors import HexColor

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from django.conf import settings
from django.utils import timezone
from datetime import datetime
import os
import tempfile
import logging

from .services import ESGDataAggregator, ReportContentGenerator

logger = logging.getLogger(__name__)


def generate_report_pdf(report):
    """
    Generate PDF report using ReportLab or serve pre-made template files
    """
    try:
        # DEBUG: Log which template type is being generated
        logger.info(f"🎯 Generating PDF for template type: {report.template.report_type}")
        logger.info(f"🎯 Report ID: {report.id}, Name: {report.name}")
        
        # Check if we have pre-made template files first
        media_root = getattr(settings, 'MEDIA_ROOT', '/mnt/c/Users/20100/v3/backend/media')
        template_files = {
            'esg_comprehensive': os.path.join(media_root, 'generated_reports', 'Enhanced_ESG_Comprehensive_Report.pdf'),
            'dst_compliance': os.path.join(media_root, 'generated_reports', 'DST_Compliance_Assessment_Report.pdf'),
            'annual_report': os.path.join(media_root, 'generated_reports', 'Enhanced_ESG_Comprehensive_Report.pdf'),  # Use ESG template for annual reports
        }
        
        # DISABLED: Use pre-made template files (forcing dynamic generation with real data)
        # template_file_path = template_files.get(report.template.report_type)
        # if template_file_path and os.path.exists(template_file_path):
        #     logger.info(f"✅ Using pre-made template file: {template_file_path}")
        #     
        #     # Copy the template file to the report
        #     with open(template_file_path, 'rb') as template_file:
        #         report.file.save(
        #             f"{report.name}.pdf",
        #             template_file,
        #             save=True
        #         )
        #     
        #     logger.info(f"📄 Report saved using template file for {report.template.report_type}")
        #     return report.file.path
        
        # Fallback to dynamic generation if no template file exists
        logger.info(f"⚙️ Generating dynamic content for {report.template.report_type}")
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        
        # Create PDF document
        doc = SimpleDocTemplate(
            temp_file.name,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build content
        story = []
        styles = getSampleStyleSheet()
        
        # Add content based on report template
        if report.template.report_type == 'esg_comprehensive':
            story.extend(_build_esg_comprehensive_content(report, styles))
        elif report.template.report_type == 'annual_report':
            # Annual reports use comprehensive ESG content
            story.extend(_build_esg_comprehensive_content(report, styles))
        elif report.template.report_type == 'dst_compliance':
            story.extend(_build_dst_compliance_content(report, styles))
        elif report.template.report_type == 'green_key':
            story.extend(_build_green_key_content(report, styles))
        elif report.template.report_type == 'quarterly_summary':
            story.extend(_build_quarterly_summary_content(report, styles))
        elif report.template.report_type == 'benchmark_analysis':
            # Benchmark analysis uses comprehensive data with focus on comparisons
            story.extend(_build_benchmark_analysis_content(report, styles))
        elif report.template.report_type == 'compliance_tracker':
            # Compliance tracker uses DST-style content
            story.extend(_build_dst_compliance_content(report, styles))
        else:
            story.extend(_build_default_content(report, styles))
        
        # Build PDF
        doc.build(story)
        
        # Save to report model
        with open(temp_file.name, 'rb') as f:
            report.file.save(
                f"{report.name}.pdf",
                f,
                save=True
            )
        
        # Clean up temp file
        os.unlink(temp_file.name)
        
        logger.info(f"📄 Dynamic report generated successfully for {report.template.report_type}")
        return report.file.path
        
    except Exception as e:
        logger.error(f"PDF generation failed: {e}")
        raise


def generate_report_excel(report):
    """
    Generate Excel report using openpyxl
    """
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add content based on report template
        if report.template.report_type == 'esg_comprehensive':
            _build_esg_excel_sheets(wb, report)
        elif report.template.report_type == 'dst_compliance':
            _build_dst_excel_sheets(wb, report)
        elif report.template.report_type == 'custom_export':
            _build_custom_excel_sheets(wb, report)
        else:
            _build_default_excel_sheets(wb, report)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        # Save to report model
        with open(temp_file.name, 'rb') as f:
            report.file.save(
                f"{report.name}.xlsx",
                f,
                save=True
            )
        
        # Clean up
        os.unlink(temp_file.name)
        
        return report.file.path
        
    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        raise


def _build_esg_comprehensive_content(report, styles):
    """Build ESG comprehensive report content with integrated data"""
    story = []
    company = report.company
    
    # Initialize data aggregator with unique seed per report
    aggregator = ESGDataAggregator(company, report.period_start, report.period_end)
    
    # Pass report ID to ensure completely unique data per report
    aggregator.report_id = str(report.id)
    
    # DEBUG: Log unique data generation details
    logger.info(f"🎲 Generating ESG data for report: {report.id}")
    logger.info(f"🎲 Company: {company.name}")
    logger.info(f"🎲 Period: {report.period_start} to {report.period_end}")
    
    data = aggregator.collect_comprehensive_data()
    content_generator = ReportContentGenerator(data)
    
    # DEBUG: Log key data variations for debugging
    logger.info(f"🔍 Report {report.id} Debug Data:")
    logger.info(f"  📊 ESG Score: {data['esg_scores']['overall_score']}")
    logger.info(f"  🏭 Company Size: {data.get('environmental_data', {}).get('company_size_category', 'unknown')}")
    logger.info(f"  👥 Employee Count: {data.get('social_data', {}).get('employee_metrics', {}).get('total_employees', 'unknown')}")
    logger.info(f"  ⚡ Energy kWh: {data.get('environmental_data', {}).get('energy_consumption', {}).get('total_kwh', 'unknown')}")
    logger.info(f"  🌱 Performance Level: {data.get('esg_scores', {}).get('performance_level', 'unknown')}")
    
    # Enhanced title page with brand colors
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=28,
        textColor=HexColor('#2EC57D'),
        alignment=TA_CENTER,
        spaceAfter=40,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=HexColor('#3DAEFF'),
        alignment=TA_CENTER,
        spaceAfter=30,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("ESG Comprehensive Report", title_style))
    story.append(Paragraph(f"{company.name}", subtitle_style))
    story.append(Paragraph(
        f"Report Period: {report.period_start.strftime('%B %d, %Y')} - {report.period_end.strftime('%B %d, %Y')}",
        styles['Normal']
    ))
    story.append(Spacer(1, 30))
    
    # Executive Summary with AI-generated content
    exec_heading_style = ParagraphStyle(
        'ExecHeading',
        parent=styles['Heading1'],
        textColor=HexColor('#2EC57D'),
        spaceBefore=20,
        spaceAfter=15
    )
    
    story.append(Paragraph("Executive Summary", exec_heading_style))
    story.append(Paragraph(
        content_generator.generate_executive_summary(),
        styles['Normal']
    ))
    story.append(Spacer(1, 25))
    
    # ESG Performance Overview with real data
    performance_heading_style = ParagraphStyle(
        'PerfHeading',
        parent=styles['Heading2'],
        textColor=HexColor('#2EC57D'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    story.append(Paragraph("ESG Performance Overview", performance_heading_style))
    
    esg_scores = data['esg_scores']
    benchmarks = data['benchmarks']['comparisons']
    
    score_data = [
        ['Category', 'Score', 'Industry Average', 'Top Quartile', 'Performance'],
        [
            'Environmental', 
            f"{esg_scores['environmental_score']:.0f}%", 
            f"{benchmarks['environmental']['industry_average']:.0f}%",
            f"{benchmarks['environmental']['top_quartile']:.0f}%",
            benchmarks['environmental']['performance'].replace('_', ' ').title()
        ],
        [
            'Social', 
            f"{esg_scores['social_score']:.0f}%", 
            f"{benchmarks['social']['industry_average']:.0f}%",
            f"{benchmarks['social']['top_quartile']:.0f}%",
            benchmarks['social']['performance'].replace('_', ' ').title()
        ],
        [
            'Governance', 
            f"{esg_scores['governance_score']:.0f}%", 
            f"{benchmarks['governance']['industry_average']:.0f}%",
            f"{benchmarks['governance']['top_quartile']:.0f}%",
            benchmarks['governance']['performance'].replace('_', ' ').title()
        ],
        [
            'Overall ESG', 
            f"{esg_scores['overall_score']:.0f}%", 
            f"{benchmarks['overall_esg']['industry_average']:.0f}%",
            f"{benchmarks['overall_esg']['top_quartile']:.0f}%",
            benchmarks['overall_esg']['performance'].replace('_', ' ').title()
        ]
    ]
    
    score_table = Table(score_data, colWidths=[1.8*inch, 0.8*inch, 1.2*inch, 1*inch, 1.2*inch])
    score_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2EC57D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#F8F9FA')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
        ('GRID', (0, 0), (-1, -1), 1, HexColor('#E0E0E0'))
    ]))
    
    story.append(score_table)
    story.append(Spacer(1, 25))
    
    # Key Achievements with real data
    story.append(Paragraph("Key Achievements", performance_heading_style))
    
    achievements = data['key_achievements']
    bullet_style = ParagraphStyle(
        'BulletStyle',
        parent=styles['Normal'],
        leftIndent=20,
        bulletIndent=10,
        spaceAfter=8
    )
    
    for achievement in achievements[:6]:  # Show top 6 achievements
        story.append(Paragraph(f"• {achievement}", bullet_style))
    
    story.append(Spacer(1, 25))
    
    # Priority Recommendations with AI-generated content
    story.append(Paragraph("Priority Recommendations", performance_heading_style))
    
    recommendations = data['recommendations']
    
    for i, rec in enumerate(recommendations[:4], 1):  # Show top 4 recommendations
        # Priority indicator
        priority_color = HexColor('#FF4444') if rec['priority'] == 'high' else HexColor('#FFA500') if rec['priority'] == 'medium' else HexColor('#4CAF50')
        
        rec_style = ParagraphStyle(
            f'RecStyle{i}',
            parent=styles['Normal'],
            leftIndent=15,
            spaceAfter=12,
            bulletIndent=10
        )
        
        priority_text = f"<font color='{priority_color.hexval()}'>[{rec['priority'].upper()}]</font>"
        story.append(Paragraph(
            f"{i}. {priority_text} <b>{rec['title']}</b><br/>"
            f"   {rec['description']}<br/>"
            f"   <i>Expected Impact:</i> {rec['impact']} | <i>Timeline:</i> {rec['timeline']}",
            rec_style
        ))
    
    story.append(Spacer(1, 25))
    
    # Data Quality Assessment
    story.append(Paragraph("Data Quality & Completeness", performance_heading_style))
    
    quality_data = data['data_quality']
    quality_text = f"""
    Overall data completeness: <b>{quality_data['overall_completeness']:.0f}%</b><br/>
    Environmental data: {quality_data['environmental_completeness']:.0f}% | 
    Social data: {quality_data['social_completeness']:.0f}% | 
    Governance data: {quality_data['governance_completeness']:.0f}%<br/>
    Evidence documentation: {quality_data['evidence_completeness']:.0f}%<br/>
    Data quality score: <b>{quality_data['quality_score']:.0f}%</b>
    """
    
    story.append(Paragraph(quality_text, styles['Normal']))
    story.append(Spacer(1, 15))
    
    # Compliance Summary
    story.append(Paragraph("Compliance Status", performance_heading_style))
    story.append(Paragraph(
        content_generator.generate_compliance_summary(),
        styles['Normal']
    ))
    
    return story


def _build_dst_compliance_content(report, styles):
    """Build Dubai Sustainable Tourism compliance report content with real data"""
    story = []
    company = report.company
    
    # Initialize DST data aggregator
    aggregator = ESGDataAggregator(company, report.period_start, report.period_end)
    data = aggregator.collect_dst_compliance_data()
    
    # Enhanced title with DST branding
    title_style = ParagraphStyle(
        'DSTTitle',
        parent=styles['Title'],
        fontSize=26,
        textColor=HexColor('#3DAEFF'),
        alignment=TA_CENTER,
        spaceAfter=30,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("Dubai Sustainable Tourism Compliance Report", title_style))
    story.append(Paragraph(f"{company.name}", styles['Heading1']))
    story.append(Spacer(1, 30))
    
    # Compliance Status with real data
    compliance_heading_style = ParagraphStyle(
        'ComplianceHeading',
        parent=styles['Heading2'],
        textColor=HexColor('#3DAEFF'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    story.append(Paragraph("Compliance Status", compliance_heading_style))
    
    compliance_score = data['compliance_score']
    readiness = data['certification_readiness']
    
    status_color = HexColor('#4CAF50') if compliance_score >= 85 else HexColor('#FFA500') if compliance_score >= 70 else HexColor('#FF4444')
    status_text = 'Compliant' if compliance_score >= 85 else 'Near Compliant' if compliance_score >= 70 else 'Non-Compliant'
    
    story.append(Paragraph(
        f"Current compliance level: <font color='{status_color.hexval()}'><b>{compliance_score:.1f}%</b></font><br/>"
        f"Status: <font color='{status_color.hexval()}'><b>{status_text}</b></font><br/>"
        f"Certification readiness: <b>{readiness['overall_readiness'].replace('_', ' ').title()}</b><br/>"
        f"Estimated certification date: {readiness['estimated_certification_date']}",
        styles['Normal']
    ))
    story.append(Spacer(1, 25))
    
    # Requirements checklist with real data
    story.append(Paragraph("DST Requirements Checklist", compliance_heading_style))
    
    requirements = data['dst_requirements']
    requirements_data = [['Requirement', 'Status', 'Score', 'Evidence']]
    
    for req in requirements:
        status_color = '#4CAF50' if req['status'] == 'complete' else '#FFA500' if req['status'] == 'in_progress' else '#FF4444'
        evidence_status = '✅' if req['evidence_uploaded'] else '❌'
        
        requirements_data.append([
            req['requirement'],
            req['status'].replace('_', ' ').title(),
            f"{req['score']:.0f}%",
            evidence_status
        ])
    
    req_table = Table(requirements_data, colWidths=[2.5*inch, 1.2*inch, 0.8*inch, 0.5*inch])
    req_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#3DAEFF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#F8F9FA')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
        ('GRID', (0, 0), (-1, -1), 1, HexColor('#E0E0E0'))
    ]))
    
    story.append(req_table)
    story.append(Spacer(1, 25))
    
    # Gap Analysis
    story.append(Paragraph("Gap Analysis & Action Plan", compliance_heading_style))
    
    gap_analysis = data['gap_analysis']
    action_plan = data['action_plan']
    
    gap_text = f"""
    <b>Identified Gaps:</b> {gap_analysis['total_gaps']} requirements need attention<br/>
    <b>Critical gaps:</b> {len(gap_analysis['critical_gaps'])} high-priority items<br/>
    <b>Missing evidence:</b> {len(gap_analysis['missing_evidence'])} documentation items<br/>
    <b>Estimated completion:</b> {gap_analysis['estimated_completion_time']}
    """
    
    story.append(Paragraph(gap_text, styles['Normal']))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Priority Actions:", styles['Heading3']))
    
    for action in action_plan:
        priority_color = HexColor('#FF4444') if action['priority'] == 'high' else HexColor('#FFA500')
        action_style = ParagraphStyle(
            'ActionStyle',
            parent=styles['Normal'],
            leftIndent=15,
            spaceAfter=10
        )
        
        story.append(Paragraph(
            f"<font color='{priority_color.hexval()}'>[{action['priority'].upper()}]</font> "
            f"<b>{action['action']}</b><br/>"
            f"Responsible: {action['responsible']} | Timeline: {action['timeline']}<br/>"
            f"Resources: {action['resources_needed']}",
            action_style
        ))
    
    return story


def _build_green_key_content(report, styles):
    """Build Green Key certification report content with real data"""
    story = []
    company = report.company
    
    # Initialize Green Key data aggregator
    aggregator = ESGDataAggregator(company, report.period_start, report.period_end)
    data = aggregator.collect_green_key_data()
    
    # Green Key branded title
    title_style = ParagraphStyle(
        'GreenKeyTitle',
        parent=styles['Title'],
        fontSize=26,
        textColor=HexColor('#4CAF50'),
        alignment=TA_CENTER,
        spaceAfter=30,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("Green Key Certification Assessment", title_style))
    story.append(Paragraph(f"{company.name}", styles['Heading1']))
    story.append(Spacer(1, 30))
    
    # Certification status with real data
    green_heading_style = ParagraphStyle(
        'GreenHeading',
        parent=styles['Heading2'],
        textColor=HexColor('#4CAF50'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    story.append(Paragraph("Certification Progress", green_heading_style))
    
    progress = data['certification_progress']
    timeline = data['implementation_timeline']
    
    progress_color = HexColor('#4CAF50') if progress >= 80 else HexColor('#FFA500') if progress >= 60 else HexColor('#FF4444')
    
    story.append(Paragraph(
        f"Current progress: <font color='{progress_color.hexval()}'><b>{progress:.0f}%</b></font><br/>"
        f"Status: <b>{'Certification Ready' if progress >= 80 else 'In Progress'}</b><br/>"
        f"Target completion: {timeline['target_completion']}<br/>"
        f"Estimated timeline: {timeline['estimated_months']:.0f} months",
        styles['Normal']
    ))
    story.append(Spacer(1, 25))
    
    # Criteria Assessment
    story.append(Paragraph("Green Key Criteria Assessment", green_heading_style))
    
    criteria = data['green_key_criteria']
    criteria_data = [['Category', 'Criteria', 'Status', 'Score']]
    
    for criterion in criteria:
        status_color = '#4CAF50' if criterion['status'] == 'complete' else '#FFA500'
        criteria_data.append([
            criterion['category'],
            criterion['criteria'],
            criterion['status'].replace('_', ' ').title(),
            f"{criterion['score']:.0f}%"
        ])
    
    criteria_table = Table(criteria_data, colWidths=[1.5*inch, 2*inch, 1.2*inch, 0.8*inch])
    criteria_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#F8F9FA')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
        ('GRID', (0, 0), (-1, -1), 1, HexColor('#E0E0E0'))
    ]))
    
    story.append(criteria_table)
    story.append(Spacer(1, 25))
    
    # Missing Requirements
    missing_requirements = data['missing_requirements']
    if missing_requirements:
        story.append(Paragraph("Outstanding Requirements", green_heading_style))
        
        bullet_style = ParagraphStyle(
            'GreenBulletStyle',
            parent=styles['Normal'],
            leftIndent=20,
            bulletIndent=10,
            spaceAfter=8
        )
        
        for requirement in missing_requirements:
            story.append(Paragraph(f"• {requirement}", bullet_style))
    
    return story


def _build_quarterly_summary_content(report, styles):
    """Build quarterly summary report content with trend analysis"""
    story = []
    company = report.company
    
    # Initialize data aggregator with unique seed per report
    aggregator = ESGDataAggregator(company, report.period_start, report.period_end)
    
    # Pass report ID to ensure completely unique data per report
    aggregator.report_id = str(report.id)
    
    # DEBUG: Log unique data generation details
    logger.info(f"🎲 Generating Quarterly data for report: {report.id}")
    logger.info(f"🎲 Company: {company.name}")
    logger.info(f"🎲 Period: {report.period_start} to {report.period_end}")
    
    data = aggregator.collect_comprehensive_data()
    
    # Quarterly branded title
    title_style = ParagraphStyle(
        'QuarterlyTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=HexColor('#2EC57D'),
        alignment=TA_CENTER,
        spaceAfter=25,
        fontName='Helvetica-Bold'
    )
    
    quarter = f"Q{((report.period_end.month-1)//3)+1} {report.period_end.year}"
    story.append(Paragraph("Quarterly ESG Summary", title_style))
    story.append(Paragraph(f"{company.name} - {quarter}", styles['Heading1']))
    story.append(Spacer(1, 25))
    
    # Performance Summary
    quarterly_heading_style = ParagraphStyle(
        'QuarterlyHeading',
        parent=styles['Heading2'],
        textColor=HexColor('#2EC57D'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    story.append(Paragraph("Quarter Performance Summary", quarterly_heading_style))
    
    trends = data['trends']
    esg_scores = data['esg_scores']
    
    # Create performance summary with trend indicators
    trend_icons = {
        'improving': '↗️',
        'stable': '➡️', 
        'declining': '↘️'
    }
    
    summary_text = f"""
    <b>Overall ESG Score:</b> {esg_scores['overall_score']:.0f}% {trend_icons.get(trends['esg_score_trend']['direction'], '')}<br/>
    <b>Environmental:</b> {esg_scores['environmental_score']:.0f}% {trend_icons.get(trends['environmental_trend']['direction'], '')}<br/>
    <b>Social:</b> {esg_scores['social_score']:.0f}% {trend_icons.get(trends['social_trend']['direction'], '')}<br/>
    <b>Governance:</b> {esg_scores['governance_score']:.0f}% {trend_icons.get(trends['governance_trend']['direction'], '')}<br/><br/>
    <b>Data Completion:</b> {esg_scores['data_completion']:.0f}% | <b>Evidence Completion:</b> {esg_scores['evidence_completion']:.0f}%
    """
    
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Key Achievements (Top 4 for quarterly)
    story.append(Paragraph("Quarter Highlights", quarterly_heading_style))
    
    achievements = data['key_achievements'][:4]
    bullet_style = ParagraphStyle(
        'QuarterlyBulletStyle',
        parent=styles['Normal'],
        leftIndent=20,
        bulletIndent=10,
        spaceAfter=8
    )
    
    for achievement in achievements:
        story.append(Paragraph(f"• {achievement}", bullet_style))
    
    story.append(Spacer(1, 20))
    
    # Next Quarter Priorities
    story.append(Paragraph("Next Quarter Priorities", quarterly_heading_style))
    
    recommendations = data['recommendations'][:3]  # Top 3 for next quarter
    for i, rec in enumerate(recommendations, 1):
        priority_color = HexColor('#FF4444') if rec['priority'] == 'high' else HexColor('#FFA500')
        
        story.append(Paragraph(
            f"{i}. <font color='{priority_color.hexval()}'>[{rec['priority'].upper()}]</font> "
            f"<b>{rec['title']}</b><br/>"
            f"   Timeline: {rec['timeline']} | Impact: {rec['impact']}",
            bullet_style
        ))
    
    return story


def _build_benchmark_analysis_content(report, styles):
    """Build benchmark analysis report content with focus on industry comparisons"""
    story = []
    company = report.company
    
    # Initialize data aggregator with unique seed per report
    aggregator = ESGDataAggregator(company, report.period_start, report.period_end)
    aggregator.report_id = str(report.id)
    
    logger.info(f"🎲 Generating Benchmark data for report: {report.id}")
    logger.info(f"🎲 Company: {company.name}")
    
    data = aggregator.collect_comprehensive_data()
    
    # Benchmark branded title
    title_style = ParagraphStyle(
        'BenchmarkTitle',
        parent=styles['Title'],
        fontSize=26,
        textColor=HexColor('#3DAEFF'),
        alignment=TA_CENTER,
        spaceAfter=30,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("ESG Benchmark Analysis", title_style))
    story.append(Paragraph(f"{company.name}", styles['Heading1']))
    story.append(Spacer(1, 30))
    
    # Benchmark comparison focus
    benchmark_heading_style = ParagraphStyle(
        'BenchmarkHeading',
        parent=styles['Heading2'],
        textColor=HexColor('#3DAEFF'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    story.append(Paragraph("Industry Benchmark Performance", benchmark_heading_style))
    
    benchmarks = data['benchmarks']['comparisons']
    esg_scores = data['esg_scores']
    
    # Create detailed benchmark table
    benchmark_data = [
        ['ESG Category', 'Your Score', 'Industry Avg', 'Top Quartile', 'Performance Gap', 'Industry Rank']
    ]
    
    categories = ['environmental', 'social', 'governance', 'overall_esg']
    category_names = ['Environmental', 'Social', 'Governance', 'Overall ESG']
    scores = [esg_scores['environmental_score'], esg_scores['social_score'], 
              esg_scores['governance_score'], esg_scores['overall_score']]
    
    for i, (cat, name, score) in enumerate(zip(categories, category_names, scores)):
        bench = benchmarks[cat]
        gap = score - bench['industry_average']
        rank = f"{bench.get('industry_rank', 'N/A')}/100"
        
        benchmark_data.append([
            name,
            f"{score:.0f}%",
            f"{bench['industry_average']:.0f}%",
            f"{bench['top_quartile']:.0f}%",
            f"{gap:+.0f}%",
            rank
        ])
    
    benchmark_table = Table(benchmark_data, colWidths=[1.5*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 0.7*inch])
    benchmark_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#3DAEFF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#F8F9FA')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
        ('GRID', (0, 0), (-1, -1), 1, HexColor('#E0E0E0'))
    ]))
    
    story.append(benchmark_table)
    story.append(Spacer(1, 25))
    
    # Key insights from benchmarking
    story.append(Paragraph("Benchmark Insights", benchmark_heading_style))
    
    insights = [
        f"Your overall ESG score of {esg_scores['overall_score']:.0f}% ranks in the {benchmarks['overall_esg']['performance'].replace('_', ' ')} tier",
        f"Strongest performance: {max(category_names[:3], key=lambda x: scores[category_names.index(x)])} ({max(scores[:3]):.0f}%)",
        f"Greatest opportunity: {min(category_names[:3], key=lambda x: scores[category_names.index(x)])} ({min(scores[:3]):.0f}%)",
        f"Industry position: {benchmarks['overall_esg'].get('percentile', 'N/A')} percentile across {company.business_sector} sector"
    ]
    
    bullet_style = ParagraphStyle(
        'BenchmarkBulletStyle',
        parent=styles['Normal'],
        leftIndent=20,
        bulletIndent=10,
        spaceAfter=8
    )
    
    for insight in insights:
        story.append(Paragraph(f"• {insight}", bullet_style))
    
    return story


def _build_default_content(report, styles):
    """Build default report content"""
    story = []
    company = report.company
    
    story.append(Paragraph(f"{report.template.display_name}", styles['Title']))
    story.append(Paragraph(f"{company.name}", styles['Heading1']))
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("Report Summary", styles['Heading2']))
    story.append(Paragraph(
        f"This report was generated on {timezone.now().strftime('%B %d, %Y')} "
        f"covering the period from {report.period_start} to {report.period_end}.",
        styles['Normal']
    ))
    
    return story


def _build_esg_excel_sheets(wb, report):
    """Build ESG comprehensive Excel sheets"""
    company = report.company
    
    # Summary sheet
    ws_summary = wb.create_sheet(title="ESG Summary")
    
    # Header
    ws_summary['A1'] = "ESG Comprehensive Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="2EC57D")
    ws_summary['A2'] = company.name
    ws_summary['A2'].font = Font(size=14, bold=True)
    
    # ESG Scores
    ws_summary['A4'] = "ESG Category"
    ws_summary['B4'] = "Score"
    ws_summary['C4'] = "Industry Average"
    
    scores = [
        ("Environmental", company.environmental_score or 72, 68),
        ("Social", company.social_score or 78, 74),
        ("Governance", company.governance_score or 75, 71),
        ("Overall ESG", company.overall_esg_score or 75, 71)
    ]
    
    for i, (category, score, avg) in enumerate(scores, 5):
        ws_summary[f'A{i}'] = category
        ws_summary[f'B{i}'] = f"{score}%"
        ws_summary[f'C{i}'] = f"{avg}%"
    
    # Environmental sheet
    ws_env = wb.create_sheet(title="Environmental")
    ws_env['A1'] = "Environmental Metrics"
    ws_env['A1'].font = Font(size=14, bold=True)
    
    # Social sheet
    ws_social = wb.create_sheet(title="Social")
    ws_social['A1'] = "Social Metrics"
    ws_social['A1'].font = Font(size=14, bold=True)
    
    # Governance sheet
    ws_gov = wb.create_sheet(title="Governance")
    ws_gov['A1'] = "Governance Metrics"
    ws_gov['A1'].font = Font(size=14, bold=True)


def _build_dst_excel_sheets(wb, report):
    """Build DST compliance Excel sheets"""
    company = report.company
    
    ws = wb.create_sheet(title="DST Compliance")
    ws['A1'] = "Dubai Sustainable Tourism Compliance"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = company.name
    
    # Compliance checklist
    ws['A4'] = "Requirement"
    ws['B4'] = "Status"
    ws['C4'] = "Score"
    
    requirements = [
        ("Environmental Management System", "Complete", "95%"),
        ("Energy Efficiency Measures", "Complete", "88%"),
        ("Water Conservation Program", "In Progress", "75%"),
        ("Waste Reduction Initiative", "Complete", "92%"),
    ]
    
    for i, (req, status, score) in enumerate(requirements, 5):
        ws[f'A{i}'] = req
        ws[f'B{i}'] = status
        ws[f'C{i}'] = score


def _build_custom_excel_sheets(wb, report):
    """Build custom export Excel sheets"""
    company = report.company
    config = report.parameters
    
    ws = wb.create_sheet(title="Custom Export")
    ws['A1'] = config.get('name', 'Custom ESG Export')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = company.name
    
    # Add selected categories
    row = 4
    for category in config.get('include_categories', []):
        ws[f'A{row}'] = f"{category.title()} Metrics"
        ws[f'A{row}'].font = Font(bold=True)
        row += 2


def _build_default_excel_sheets(wb, report):
    """Build default Excel sheets"""
    company = report.company
    
    ws = wb.create_sheet(title="Report")
    ws['A1'] = report.template.display_name
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = company.name
    ws['A4'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"